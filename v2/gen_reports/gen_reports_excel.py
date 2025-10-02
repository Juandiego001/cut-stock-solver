import random
from classes import Solution
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment



general_alignment = Alignment(horizontal='center', vertical='center')
black_border = Border(left=Side(border_style='thin', color='FF000000'),
                      right=Side(border_style='thin', color='FF000000'),
                      top=Side(border_style='thin', color='FF000000'),
                      bottom=Side(border_style='thin', color='FF000000'))
white_border = Border(left=Side(border_style='thin', color='00FFFFFF'),
                      right=Side(border_style='thin', color='00FFFFFF'),
                      top=Side(border_style='thin', color='00FFFFFF'),
                      bottom=Side(border_style='thin', color='00FFFFFF'))
headers_font = Font(color='FFFFFF', bold=True)
headers_bg = PatternFill(fill_type='solid', fgColor='00023047')

colors = {}
saved_colors = []


def write_test_case(ws, case_data):
    '''Escribir hoja de excel de caso de prueba'''

    # Headers
    headers = ['A1', 'B1', 'A3', 'B3', 'C3', 'D3']
    for header in headers:
        ws[header].font = headers_font
        ws[header].fill = headers_bg
        ws[header].alignment = general_alignment
        ws[header].border = white_border

    # Change column widths
    cols = ['A', 'B']
    for col in cols:
        ws.column_dimensions[col].width = 14

    # Ancho y largo grande
    ws['A1'] = 'Ancho grande'
    ws['A2'] = case_data['ancho_grande']
    ws['A2'].border = black_border
    ws['A2'].alignment = general_alignment

    ws['B1'] = 'Largo grande'
    ws['B2'] = case_data['largo_grande']
    ws['B2'].border = black_border
    ws['B2'].alignment = general_alignment

    # Items
    ws['A3'] = 'Id'
    ws['B3'] = 'Demanda'
    ws['C3'] = 'Ancho'
    ws['D3'] = 'Largo'

    for i, item in enumerate(case_data['items']):
        ws[f'A{i+4}'] = item.id
        ws[f'B{i+4}'] = item.dem
        ws[f'C{i+4}'] = item.ancho
        ws[f'D{i+4}'] = item.largo

        ws[f'A{i+4}'].border = black_border
        ws[f'B{i+4}'].border = black_border
        ws[f'C{i+4}'].border = black_border
        ws[f'D{i+4}'].border = black_border

        ws[f'A{i+4}'].alignment = general_alignment
        ws[f'B{i+4}'].alignment = general_alignment
        ws[f'C{i+4}'].alignment = general_alignment
        ws[f'D{i+4}'].alignment = general_alignment


def write_iterations_summary_header(ws):
    '''Escribir la cabecera del resumen de las iteraciones'''

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}1'].font = headers_font
        ws[f'{col}1'].fill = headers_bg
        ws[f'{col}1'].alignment = general_alignment
        ws[f'{col}1'].border = white_border
        ws.column_dimensions[col].width = 14
        ws.row_dimensions[1].height = 24

    # Change column widths
    cols = ['B', 'C', 'D', 'E']
    for col in cols:
        ws.column_dimensions[col].width = 23

    ws['A1'] = 'Iteración'
    ws['B1'] = 'Mejor Fitness Swap'
    ws['C1'] = 'Mejor Fitness Insertions'
    ws['D1'] = 'Mejor Fitness 2OPT'
    ws['E1'] = 'Mejor Fitness Actual'
    ws['F1'] = 'Mudarse a'


def create_sheets_summary():
    '''Crea las hojas del excel summary'''

    wb = Workbook()

    wb.active.title = 'Caso'
    wb.create_sheet('SolInicial_Info')
    wb.create_sheet('SolInicial_Debug')
    wb.create_sheet('Iteraciones')
    wb.create_sheet('Solucion_Info')
    wb.create_sheet('Solucion_Debug')

    write_iterations_summary_header(wb['Iteraciones'])

    return wb


def write_iterations_summary(ws, iteration, z_swap, z_insertions, z_2opt, z_best):
    '''Escribir hoja de iteraciones resumida'''

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{iteration+1}'].border = black_border
        ws[f'{col}{iteration+1}'].alignment = general_alignment

    ws[f'A{iteration+1}'] = iteration
    ws[f'B{iteration+1}'] = z_swap
    ws[f'C{iteration+1}'] = z_insertions
    ws[f'D{iteration+1}'] = z_2opt
    ws[f'E{iteration+1}'] = z_best

    if z_swap < z_insertions and z_swap < z_2opt and z_swap < z_best:
        ws[f'B{iteration+1}'].font = Font(color='FFFFFF', bold=True)
        ws[f'B{iteration+1}'].fill = PatternFill(
            fill_type='solid', fgColor='00963634')
        ws[f'F{iteration+1}'] = 'Vecino Swap'
    elif z_insertions < z_swap and z_insertions < z_2opt and z_insertions < z_best:
        ws[f'C{iteration+1}'].font = Font(color='FFFFFF', bold=True)
        ws[f'C{iteration+1}'].fill = PatternFill(
            fill_type='solid', fgColor='00963634')
        ws[f'F{iteration+1}'] = 'Vecino Insertions'
    elif z_2opt < z_swap and z_2opt < z_insertions and z_2opt < z_best:
        ws[f'D{iteration+1}'].font = Font(color='FFFFFF', bold=True)
        ws[f'D{iteration+1}'].fill = PatternFill(
            fill_type='solid', fgColor='00963634')
        ws[f'F{iteration+1}'] = 'Vecino 2OPT'
    elif z_swap != z_best:
        ws[f'B{iteration+1}'].font = Font(color='FFFFFF', bold=True)
        ws[f'B{iteration+1}'].fill = PatternFill(
            fill_type='solid', fgColor='00963634')
        ws[f'F{iteration+1}'] = 'Vecino Swap'


def create_sheets_iterations(iteration: int):
    '''Crea las hojas de los excels de iteraciones'''

    wb = Workbook()
    wb.active.title = 'Info'
    wb.create_sheet('Debug')

    return wb


def generate_random_color():
    '''Función para generar un color aleatorio'''

    rojo = random.randint(0, 255)
    verde = random.randint(0, 255)
    azul = random.randint(0, 255)

    the_color = f'{rojo:02x}{verde:02x}{azul:02x}'

    while (the_color in saved_colors):
        rojo = random.randint(0, 255)
        verde = random.randint(0, 255)
        azul = random.randint(0, 255)

        the_color = f'{rojo:02x}{verde:02x}{azul:02x}'

    return the_color


def write_solution_iteration_info(ws: Worksheet, solutions: list[Solution]):
    '''Write solution iteration info'''

    # Vecinos title
    ws['A1'] = 'Vecinos'
    ws['A1'].font = headers_font
    ws['A1'].fill = headers_bg
    ws['A1'].alignment = general_alignment
    ws['A1'].border = white_border
    ws.column_dimensions['A'].width = 15

    permutation_size = len(solutions[0].permutation)
    desperdicio_column = permutation_size + 1
    fitness_column = permutation_size + 2
    for i, solution in enumerate(solutions):
        ws[f'A{i+2}'] = i+1
        ws[f'A{i+2}'].alignment = general_alignment
        ws[f'A{i+2}'].border = black_border

        for j, item in enumerate(solution.permutation):
            ws.cell(row=i+2, column=j+2).value = item.id
            ws.cell(row=i+2, column=j+2).alignment = general_alignment
            ws.cell(row=i+2, column=j+2).border = black_border

        ws.cell(row=i+2, column=desperdicio_column).value = solution.desperdicio
        ws.cell(row=i+2, column=desperdicio_column).alignment = general_alignment
        ws.cell(row=i+2, column=desperdicio_column).border = black_border

        ws.cell(row=i+2, column=fitness_column).value = solution.fitness
        ws.cell(row=i+2, column=fitness_column).alignment = general_alignment
        ws.cell(row=i+2, column=fitness_column).border = black_border
    
    for i in range(permutation_size):
        ws.column_dimensions[get_column_letter(i+2)].width = 5

    # Desperdicio title
    ws.cell(row=1, column=desperdicio_column).value = 'Desperdicio'
    ws.cell(row=1, column=desperdicio_column).font = headers_font
    ws.cell(row=1, column=desperdicio_column).fill = headers_bg
    ws.cell(row=1, column=desperdicio_column).alignment = general_alignment
    ws.cell(row=1, column=desperdicio_column).border = white_border
    ws.column_dimensions[get_column_letter(desperdicio_column)].width = 15

    # Fitness title
    ws.cell(row=1, column=fitness_column).value = 'Fitness'
    ws.cell(row=1, column=fitness_column).font = headers_font
    ws.cell(row=1, column=fitness_column).fill = headers_bg
    ws.cell(row=1, column=fitness_column).alignment = general_alignment
    ws.cell(row=1, column=fitness_column).border = white_border
    ws.column_dimensions[get_column_letter(fitness_column)].width = 15



def write_solution_info(ws, solution: Solution):
    '''Escribir información de solución en excel'''

    # Headers
    headers = ['A1', 'A3', 'B3', 'C3', 'D3', 'E3', 'F3', 'G1', 'H1']
    for header in headers:
        ws[header].font = headers_font
        ws[header].fill = headers_bg
        ws[header].alignment = general_alignment
        ws[header].border = white_border

    # Change column widths
    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    for col in cols:
        ws.column_dimensions[col].width = 15

    # Permutación
    ws['A1'] = 'Permutación'
    ws['A3'] = 'ID'
    ws['B3'] = 'Demanda'
    ws['C3'] = 'Ancho'
    ws['D3'] = 'Largo'
    ws['E3'] = 'Completada'
    ws['F3'] = 'Faltante'

    ws.merge_cells('A1:F2')

    i = 4
    for items in solution.permutation:
        ws[f'A{i}'] = items.id
        ws[f'B{i}'] = items.dem
        ws[f'C{i}'] = items.ancho
        ws[f'D{i}'] = items.largo
        ws[f'E{i}'] = solution.dem_com[items.id]
        ws[f'F{i}'] = solution.dem_fal[items.id]

        ws[f'A{i}'].border = black_border
        ws[f'B{i}'].border = black_border
        ws[f'C{i}'].border = black_border
        ws[f'D{i}'].border = black_border
        ws[f'E{i}'].border = black_border
        ws[f'F{i}'].border = black_border

        i += 1

    # Desperdicio
    ws['G1'] = 'Desperdicio'
    ws.merge_cells('G1:G3')
    ws['G4'] = solution.desperdicio
    ws['G4'].border = black_border

    # Fitness
    ws['H1'] = 'Fitness'
    ws.merge_cells('H1:H3')
    ws['H4'] = solution.fitness
    ws['H4'].border = black_border

    # General alignment
    for c in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        for r in range(1, i+1):
            ws[f'{c}{r}'].alignment = general_alignment


def write_solution_debug(ws, matrixes):
    '''Escribir el debug de la solución como tal'''

    # Change column widths
    base_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L',
                 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    all_cols = []
    for col1 in base_cols:
        ws.column_dimensions[col1].width = 4
        all_cols.append(col1)

    for col1 in base_cols:
        for col2 in base_cols:
            ws.column_dimensions[f'{col1}{col2}'].width = 4
            all_cols.append(f'{col1}{col2}')

    all_cols_size = len(all_cols)
    start_row_row = 2
    for i in matrixes:
        max_row = 0
        last_col_matrix = start_col_matrix = 'B'
        for matrix in i:
            start_row_matrix = start_row_row
            for k in matrix:
                cell = [start_col_matrix, start_row_matrix]

                for l in k:
                    ws[f'{cell[0]}{cell[1]}'].border = black_border

                    if l != 0:
                        if not (l in colors):
                            colors[l] = generate_random_color()

                        ws[f'{cell[0]}{cell[1]}'].fill = PatternFill(
                            fill_type='solid', fgColor=f'{colors[l]}')

                    col_pos = all_cols.index(cell[0])
                    next_col_pos = col_pos + 1 if col_pos + 1 != all_cols_size else 0
                    next_col = all_cols[next_col_pos]
                    cell = [next_col, cell[1]]
                    last_col_matrix = next_col

                start_row_matrix += 1

                max_row = start_row_matrix + 1 if start_row_matrix > max_row else max_row

            col_pos = all_cols.index(last_col_matrix)
            next_col_pos = col_pos + 1 if col_pos + 1 < all_cols_size else 0
            next_col = all_cols[next_col_pos]
            start_col_matrix = next_col

        start_row_row = max_row + 2
