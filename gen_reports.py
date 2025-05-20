from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

general_alignment = Alignment(horizontal='center', vertical='center')
headers_font = Font(color='FFFFFF', bold=True)
headers_bg = PatternFill(fill_type='solid', fgColor='00023047')
white_border = Border(left=Side(border_style='thin', color='00FFFFFF'),
                      right=Side(border_style='thin', color='00FFFFFF'),
                      top=Side(border_style='thin', color='00FFFFFF'),
                      bottom=Side(border_style='thin', color='00FFFFFF'))
black_border = Border(left=Side(border_style='thin', color='FF000000'),
                      right=Side(border_style='thin', color='FF000000'),
                      top=Side(border_style='thin', color='FF000000'),
                      bottom=Side(border_style='thin', color='FF000000'))


def create_sheets_summary():
    '''Crea las hojas del excel summary'''

    wb = Workbook()

    wb.active.title = 'Caso'
    wb.create_sheet('SolInicial')
    wb.create_sheet('Iteraciones')
    wb.create_sheet('Solucion')

    write_iterations_summary_header(wb['Iteraciones'])

    return wb


def create_sheets_iterations(iteration: int):
    '''Crea las hojas de los excels de iteraciones'''

    wb = Workbook()
    wb.active.title = f'{iteration}_VecinosT'
    wb.create_sheet(f'{iteration}_VecinosH')

    return wb


def write_test_case(ws, case_data):
    '''Escribir hoja de excel de caso de prueba'''

    # Headers
    headers = ['A1', 'B1', 'A3', 'B3', 'C3', 'D3']
    for header in headers:
        ws[header].font = headers_font
        ws[header].fill = headers_bg
        ws[header].alignment = general_alignment
        ws[header].border = white_border

    # Change column widths
    cols = ['A', 'B']
    for col in cols:
        ws.column_dimensions[col].width = 14

    # Ancho y largo grande
    ws['A1'] = 'Ancho grande'
    ws['A2'] = case_data['ancho_grande']
    ws['A2'].border = black_border
    ws['A2'].alignment = general_alignment

    ws['B1'] = 'Largo grande'
    ws['B2'] = case_data['largo_grande']
    ws['B2'].border = black_border
    ws['B2'].alignment = general_alignment

    # Items
    ws['A3'] = 'Id'
    ws['B3'] = 'Demanda'
    ws['C3'] = 'Ancho'
    ws['D3'] = 'Largo'

    for i, item in enumerate(case_data['items']):
        ws[f'A{i+4}'] = item.id
        ws[f'B{i+4}'] = item.dem
        ws[f'C{i+4}'] = item.ancho
        ws[f'D{i+4}'] = item.largo

        ws[f'A{i+4}'].border = black_border
        ws[f'B{i+4}'].border = black_border
        ws[f'C{i+4}'].border = black_border
        ws[f'D{i+4}'].border = black_border

        ws[f'A{i+4}'].alignment = general_alignment
        ws[f'B{i+4}'].alignment = general_alignment
        ws[f'C{i+4}'].alignment = general_alignment
        ws[f'D{i+4}'].alignment = general_alignment


def write_solution(ws, solution, start_row):
    '''Escribir solución en excel'''

    # Headers
    headers = [f'A{start_row}',
        f'B{start_row}', 
        f'C{start_row}', 
        f'C{start_row+2}', 
        f'D{start_row+2}', 
        f'E{start_row+2}', 
        f'F{start_row}',
        f'G{start_row}', 
        f'H{start_row}', 
        f'H{start_row+1}', 
        f'I{start_row+1}', 
        f'J{start_row+1}', 
        f'K{start_row+1}', 
        f'L{start_row+1}', 
        f'L{start_row+2}', 
        f'M{start_row+2}']
    for header in headers:
        ws[header].font = headers_font
        ws[header].fill = headers_bg
        ws[header].alignment = general_alignment
        ws[header].border = white_border

    # Change column widths
    cols = ['D', 'F', 'K', 'M']
    for col in cols:
        ws.column_dimensions[col].width = 15

    # Vectores T y H
    ws[f'A{start_row}'] = 'Vector T'
    ws[f'B{start_row}'] = 'Vector H'

    ws.merge_cells(f'A{start_row}:A{start_row+2}')
    ws.merge_cells(f'B{start_row}:B{start_row+2}')

    # Demanda completada/faltante
    ws[f'C{start_row}'] = 'Demanda'
    ws[f'C{start_row+2}'] = 'ID'
    ws[f'D{start_row+2}'] = 'Completada'
    ws[f'E{start_row+2}'] = 'Faltante'

    ws.merge_cells(f'C{start_row}:E{start_row+1}')

    # Desperdicio
    ws[f'F{start_row}'] = 'Desperdicio'
    ws.merge_cells(f'F{start_row}:F{start_row+2}')
    ws[f'F{start_row+3}'] = solution.desperdicio
    ws[f'F{start_row+3}'].border = black_border

    # Fitness
    ws[f'G{start_row}'] = 'Fitness'
    ws.merge_cells(f'G{start_row}:G{start_row+2}')
    ws[f'G{start_row+3}'] = solution.fitness
    ws[f'G{start_row+3}'].border = black_border

    # Subespacios
    ws[f'H{start_row}'] = 'Subespacios'
    ws.merge_cells(f'H{start_row}:M{start_row}')

    ws[f'H{start_row+1}'] = '#'
    ws[f'I{start_row+1}'] = 'Ancho'
    ws[f'J{start_row+1}'] = 'Largo'
    ws[f'K{start_row+1}'] = 'Área disponible'

    ws.merge_cells(f'H{start_row+1}:H{start_row+2}')
    ws.merge_cells(f'I{start_row+1}:I{start_row+2}')
    ws.merge_cells(f'J{start_row+1}:J{start_row+2}')
    ws.merge_cells(f'K{start_row+1}:K{start_row+2}')

    ws[f'L{start_row+1}'] = 'Items'
    ws.merge_cells(f'L{start_row+1}:M{start_row+1}')

    ws[f'L{start_row+2}'] = 'ID'
    ws[f'M{start_row+2}'] = 'Capacidad'

    # Valores de T y H
    i = start_row
    for t, h in zip(solution.T, solution.H):
        ws[f'A{i+3}'] = t
        ws[f'B{i+3}'] = h

        ws[f'A{i+3}'].border = black_border
        ws[f'B{i+3}'].border = black_border

        i += 1

    # Demanda
    i = start_row
    for key in solution.dem_com:
        ws[f'C{i+3}'] = key
        ws[f'D{i+3}'] = solution.dem_com[key]
        ws[f'E{i+3}'] = solution.dem_fal[key]

        ws[f'C{i+3}'].border = black_border
        ws[f'D{i+3}'].border = black_border
        ws[f'E{i+3}'].border = black_border

        i += 1

    # Subespacios
    i = start_row
    for index, sub_es in enumerate(solution.v_sub):
        ws[f'H{i+3}'] = index + 1
        ws[f'I{i+3}'] = sub_es.ancho
        ws[f'J{i+3}'] = sub_es.largo
        ws[f'K{i+3}'] = sub_es.area_disponible

        ws[f'H{i+3}'].border = black_border
        ws[f'I{i+3}'].border = black_border
        ws[f'J{i+3}'].border = black_border
        ws[f'K{i+3}'].border = black_border

        j = i + 3
        for key in sub_es.items_capacidad:
            ws[f'L{j}'] = key
            ws[f'M{j}'] = sub_es.items_capacidad[key]

            ws[f'L{j}'].border = black_border
            ws[f'M{j}'].border = black_border

            j += 1

        ws.merge_cells(f'H{i+3}:H{j-1}')
        ws.merge_cells(f'I{i+3}:I{j-1}')
        ws.merge_cells(f'J{i+3}:J{j-1}')
        ws.merge_cells(f'K{i+3}:K{j-1}')

        i = j - 3

    # General alignment
    for c in range(1, 14):
        for r in range(start_row, i+4):
            ws.cell(row=r, column=c).alignment = general_alignment


def write_iterations(ws, vecinos, iteration, c, len_items):
    '''Escribir hoja de excel de los vecinos creados por las iteraciones'''

    start_row = 1

    # Iteración
    ws[f'A{start_row}'].font = headers_font
    ws[f'A{start_row}'].fill = headers_bg
    ws[f'A{start_row}'].alignment = general_alignment
    ws[f'A{start_row}'].border = white_border
    ws[f'A{start_row}'] = f'Iteración {iteration}'
    ws.merge_cells(f'A{start_row}:M{start_row}')

    start_row += 1
    for vecino in vecinos:
        write_solution(ws, vecino, start_row)
        start_row += ((2**c) * len_items) + 4


def write_iterations_summary_header(ws):
    '''Escribir la cabecera del resumen de las iteraciones'''

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}1'].font = headers_font
        ws[f'{col}1'].fill = headers_bg
        ws[f'{col}1'].alignment = general_alignment
        ws[f'{col}1'].border = white_border
        ws.column_dimensions[col].width = 14
        # ws.column_dimensions[col].height = 24
        ws.row_dimensions[1].height = 24

    # Change column widths
    cols = ['B', 'C', 'D']
    for col in cols:
        ws.column_dimensions[col].width = 19

    ws['A1'] = 'Iteración'
    ws['B1'] = 'Mejor Fitness T'
    ws['C1'] = 'Mejor Fitness H' 
    ws['D1'] = 'Mejor Fitness Actual'
    ws['E1'] = 'Mudarse a'


def write_iterations_summary(ws, iteration, z_T, z_H, z_best):
    '''Escribir hoja de iteraciones resumida'''

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{iteration+1}'].border = black_border
        ws[f'{col}{iteration+1}'].alignment = general_alignment

    ws[f'A{iteration+1}'] = iteration
    ws[f'B{iteration+1}'] = z_T
    ws[f'C{iteration+1}'] = z_H
    ws[f'D{iteration+1}'] = z_best    

    if z_T < z_H and z_T < z_best:
        ws[f'B{iteration+1}'].font = Font(color='FFFFFF', bold=True)
        ws[f'B{iteration+1}'].fill = PatternFill(fill_type='solid', fgColor='00963634')
        ws[f'E{iteration+1}'] = 'Vecino T'
    elif z_H < z_T and z_H < z_best:
        ws[f'C{iteration+1}'].font = Font(color='FFFFFF', bold=True)
        ws[f'C{iteration+1}'].fill = PatternFill(fill_type='solid', fgColor='00963634')
        ws[f'E{iteration+1}'] = 'Vecino H'
